"""
generate_random_games.py — generate random single-heap Heapgo games, compute
their mean and temperature, and write the results to an Excel file.

Each game is reproducible from its seed alone: the seed fixes both the number
of counters and every counter's (value, color).

Excel columns:
  1. seed           reproduces the whole game
  2. counters       number of counters (stones) in the heap
  3. heapgo_game    the heap, e.g. [(4, 'red'), (9, 'blue'), (6, 'red')]
  4. curly          CGSuite-style {L | R} form of the game tree
  5. mean           Computed via brute_force_mt
  6. temperature    Computed via brute_force_mt
  7. num_terminals  number of leaves in the game tree
  8. convert_ms     time to turn the heap into a nested-list tree (heapgo_to_tree)
  9. solve_ms       time to compute the mean and temperature (brute_force_mt) from 
                    the nested-list tree

Requires openpyxl for the Excel output.
"""

import os
import random
import time

from heapgo_to_tree import heapgo_to_tree
from brute_force import brute_force_mt
from stable_pair import ColdGameError


# ===========================================================================
# CONFIG — edit here
# ===========================================================================
BASE_SEED    = 1        # first seed; games use BASE_SEED, BASE_SEED+1, ...
NUM_GAMES    = 200      # how many games to generate
MIN_COUNTERS = 4        # fewest counters in a heap
MAX_COUNTERS = 10        # most counters in a heap
MAX_VALUE    = 9        # counter values are drawn from 1..MAX_VALUE
COLORS       = ['red', 'blue']
XLSX_NAME    = "random_heapgo_games.xlsx"

# Write the workbook next to THIS script (i.e. in the same folder as the other .py files).
XLSX_PATH    = os.path.join(os.path.dirname(os.path.abspath(__file__)), XLSX_NAME)
SHOW_EACH    = False    # print each game to the console as it is computed


# ===========================================================================
# Game generation and helpers
# ===========================================================================
def make_heap(seed):
    """Reproducibly build one random single heap from `seed`."""
    rng = random.Random(seed)
    n = rng.randint(MIN_COUNTERS, MAX_COUNTERS)
    return [(rng.randint(1, MAX_VALUE), rng.choice(COLORS)) for _ in range(n)]


def tree_to_curly(tree):
    """Nested-list game tree -> CGSuite-style {L | R} string."""
    return (str(tree)
            .replace('[', '{')
            .replace(']', '}')
            .replace(', ', ' | '))


def count_terminals(tree):
    """Number of leaves (terminals) in a nested-list game tree."""
    return str(tree).count(',') + 1


# ===========================================================================
# Main
# ===========================================================================
def generate():
    rows = []
    for i in range(NUM_GAMES):
        seed = BASE_SEED + i
        heap = make_heap(seed)

        t0 = time.perf_counter()
        tree = heapgo_to_tree(heap)                # heap is valid by construction
        convert_ms = (time.perf_counter() - t0) * 1000

        curly = tree_to_curly(tree)
        n_terminals = count_terminals(tree)

        t0 = time.perf_counter()
        try:
            mean, temp = brute_force_mt(tree)
        except ColdGameError:
            mean, temp = "cold", "cold"
        solve_ms = (time.perf_counter() - t0) * 1000

        rows.append([seed, len(heap), repr(heap), curly, mean, temp, n_terminals,
                     round(convert_ms, 4), round(solve_ms, 4)])

        if SHOW_EACH:
            print(f"seed={seed:<4} counters={len(heap)}  M={mean}  T={temp}  "
                  f"terminals={n_terminals}  convert={convert_ms:.4f}ms  "
                  f"solve={solve_ms:.4f}ms")
    return rows


def write_xlsx(rows, path=XLSX_PATH):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\nopenpyxl is not installed; skipping Excel output.")
        return

    headers = ["seed", "counters", "heapgo_game", "curly",
               "mean", "temperature", "num_terminals",
               "convert_ms", "solve_ms"]

    wb = Workbook()
    ws = wb.active
    ws.title = "heapgo games"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    # Keep the game/curly columns as literal text so Excel does not reinterpret
    # them (columns 3 and 4).
    for col in (3, 4):
        letter = get_column_letter(col)
        for cell in ws[letter]:
            cell.number_format = "@"

    wb.save(path)
    print(f"\nWrote {len(rows)} games to {path}")


def main():
    print(f"Generating {NUM_GAMES} random single-heap Heapgo games "
          f"(seeds {BASE_SEED}..{BASE_SEED + NUM_GAMES - 1}, "
          f"{MIN_COUNTERS}-{MAX_COUNTERS} counters, values 1..{MAX_VALUE})")
    print("-" * 60)
    rows = generate()
    print("-" * 60)
    write_xlsx(rows)


if __name__ == "__main__":
    main()