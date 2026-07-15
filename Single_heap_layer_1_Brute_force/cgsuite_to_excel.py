"""
cgsuite_to_excel.py  (Part 2 of 2)

Take the mean/temperature that CGSuite computed and add them to the Excel file,
so that one can compare CGSuite's answers with this project's answers side by side.

HOW TO USE
----------
Step 1 (in CGSuite):
    Run the batch block from cgsuite_input.txt in the CGSuite Worksheet
    (that block is produced by export_to_cgsuite.py). 
    Select ALL of that output, copy it, and paste it into the text file named
    below (TXT_NAME, default "cgsuite_output.txt"). Save it.

Step 2 (here):
    Just run this file:   python cgsuite_to_excel.py
    It reads the Excel file and the pasted CGSuite output, matches them by
    seed, and writes a new Excel file with these extra columns:
        cgsuite_mean, cgsuite_temp   -- CGSuite's answers
        mean_match, temp_match       -- "match"/"mismatch" vs this project's
                                        answers, coloured GREEN for a match and
                                        RED for a mismatch

Requires pandas and openpyxl.
"""

import json
import os
import re
import sys
from fractions import Fraction

import pandas as pd


# ===========================================================================
# CONFIG — edit here
# ===========================================================================
XLSX_IN_NAME  = "random_heapgo_games.xlsx"           # Excel file to read
TXT_NAME      = "cgsuite_output.txt"                  # CGSuite output pasted here
XLSX_OUT_NAME = "random_heapgo_games_with_cgsuite.xlsx"  # Excel file to write

SEED_COL      = "seed"            # seed column in the Excel file (used to match)
MEAN_COL      = "cgsuite_mean"    # new column: CGSuite's mean
TEMP_COL      = "cgsuite_temp"    # new column: CGSuite's temperature

# Compare columns: brute force value vs CGSuite's value, "match" / "mismatch".
# The cell is coloured GREEN when they match and RED when they do not.
MEAN_MATCH_COL = "mean_match"
TEMP_MATCH_COL = "temp_match"
BRUTE_FORCE_MEAN_COL   = "mean"           # this project's mean column in the Excel file
BRUTE_FORCE_TEMP_COL   = "temperature"    # this project's temperature column
GREEN          = "C6EFCE"         # fill for a match
RED            = "FFC7CE"         # fill for a mismatch

# All three files live next to THIS script.
_HERE         = os.path.dirname(os.path.abspath(__file__))
XLSX_IN_PATH  = os.path.join(_HERE, XLSX_IN_NAME)
TXT_PATH      = os.path.join(_HERE, TXT_NAME)
XLSX_OUT_PATH = os.path.join(_HERE, XLSX_OUT_NAME)


# ===========================================================================
def read_text(path):
    """Read the pasted CGSuite output, coping with whatever encoding it saved as."""
    with open(path, "rb") as f:
        raw = f.read()
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        text = raw.decode("utf-16")
    elif raw[:3] == b"\xef\xbb\xbf":
        text = raw.decode("utf-8-sig")
    elif b"\x00" in raw:                       # UTF-16 without a BOM
        try:
            text = raw.decode("utf-16-le")
        except UnicodeDecodeError:
            text = raw.decode("utf-16-be")
    else:
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("cp1252", errors="ignore")
    return text.replace("\x00", "").lstrip("\ufeff")


def parse_records(text):
    """
    Pull (seed, mean, temp) triples out of the CGSuite output.
    """
    entries = None

    # 1) a proper JSON list of "seed,mean,temp" strings
    try:
        data = json.loads(text)
        if isinstance(data, list):
            entries = [str(x) for x in data]
    except (json.JSONDecodeError, ValueError):
        pass

    # 2) not valid JSON, but the entries are still quoted "..."
    if entries is None:
        quoted = re.findall(r'"([^"]*)"', text)
        if quoted:
            entries = quoted

    # 3) last resort: one bare entry per line
    if entries is None:
        entries = [ln for ln in re.split(r"[\r\n]+", text.strip().strip("[]"))
                   if ln.strip()]

    records = {}
    for entry in entries:
        parts = [p.strip() for p in entry.split(",")]
        if len(parts) == 3:
            try:
                records[int(parts[0])] = (parts[1], parts[2])
            except ValueError:
                pass                      # not "seed,mean,temp" -> skip
    return records


def main():
    print(">>> parser version: JSON-based")
    print(f"Reading Excel     {XLSX_IN_PATH}")
    try:
        df = pd.read_excel(XLSX_IN_PATH)
    except Exception as e:
        sys.exit(f"ERROR reading Excel '{XLSX_IN_PATH}': {e}")

    if SEED_COL not in df.columns:
        sys.exit(f"ERROR: no '{SEED_COL}' column. Columns are: {list(df.columns)}")

    print(f"Reading CGSuite   {TXT_PATH}")
    try:
        text = read_text(TXT_PATH)
    except FileNotFoundError:
        sys.exit(f"ERROR: '{TXT_PATH}' not found. Paste the CGSuite output there first.")

    results = parse_records(text)
    if not results:
        sys.exit("ERROR: no 'seed,mean,temp' records found in the CGSuite output.")

    seeds = df[SEED_COL].astype("Int64")
    df[MEAN_COL] = [results.get(int(s), (None, None))[0] if pd.notna(s) else None
                    for s in seeds]
    df[TEMP_COL] = [results.get(int(s), (None, None))[1] if pd.notna(s) else None
                    for s in seeds]

    matched = int(df[MEAN_COL].notna().sum())
    missing = [int(s) for s in seeds.dropna() if int(s) not in results]

    # --- compare brute force value vs CGSuite's, exactly (2.5 == "5/2") ---
    for col in (BRUTE_FORCE_MEAN_COL, BRUTE_FORCE_TEMP_COL):
        if col not in df.columns:
            sys.exit(f"ERROR: no '{col}' column to compare against. "
                     f"Columns are: {list(df.columns)}")

    def match_col(our_col, cg_col):
        out = []
        for _, r in df.iterrows():
            if pd.isna(r[cg_col]):
                out.append(None)                     # no CGSuite value -> blank
            else:
                out.append("match" if _equal(r[our_col], r[cg_col]) else "mismatch")
        return out

    df[MEAN_MATCH_COL] = match_col(BRUTE_FORCE_MEAN_COL, MEAN_COL)
    df[TEMP_MATCH_COL] = match_col(BRUTE_FORCE_TEMP_COL, TEMP_COL)

    mean_mm = df[MEAN_MATCH_COL].eq("mismatch").sum()
    temp_mm = df[TEMP_MATCH_COL].eq("mismatch").sum()

    df.to_excel(XLSX_OUT_PATH, index=False)
    _colour_match_cells(XLSX_OUT_PATH, df)

    print(f"Parsed {len(results)} CGSuite records; matched {matched} of {len(df)} rows by seed.")
    print(f"mean:  {mean_mm} mismatch(es)   temperature:  {temp_mm} mismatch(es)")
    if missing:
        show = ", ".join(map(str, missing[:10])) + (" ..." if len(missing) > 10 else "")
        print(f"WARNING: {len(missing)} row(s) had no matching CGSuite seed: {show}")
    print(f"Wrote  {XLSX_OUT_PATH}")


def _equal(ours, theirs):
    """Exact equality via fractions, so 2.5 and '5/2' are equal."""
    try:
        return Fraction(str(ours)).limit_denominator(10**9) == Fraction(str(theirs))
    except (ValueError, ZeroDivisionError):
        return False


def _colour_match_cells(path, df):
    """Fill each match cell green and each mismatch cell red."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    green = PatternFill("solid", fgColor=GREEN)
    red   = PatternFill("solid", fgColor=RED)

    wb = load_workbook(path)
    ws = wb.active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    for name in (MEAN_MATCH_COL, TEMP_MATCH_COL):
        col = headers[name]
        for row in range(2, ws.max_row + 1):          # row 1 is the header
            value = ws.cell(row=row, column=col).value
            if value == "match":
                ws.cell(row=row, column=col).fill = green
            elif value == "mismatch":
                ws.cell(row=row, column=col).fill = red
    wb.save(path)


if __name__ == "__main__":
    main()